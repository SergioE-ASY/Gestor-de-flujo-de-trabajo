from datetime import timedelta

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.db.models import Sum
from django.utils import timezone
from .models import Organization, OrganizationUser
from .permissions import can_manage_members, can_assign_role, is_last_owner
from shared.decorators import require_org_member, org_permission
from projects.models import Project

NAME_CHANGE_DAYS = 90

User = get_user_model()


@login_required
def dashboard(request):
    user_orgs = OrganizationUser.objects.filter(user=request.user).select_related('organization')

    user_projects = Project.objects.filter(
        members__user=request.user, deleted_at__isnull=True
    ).select_related('organization', 'owner').distinct()[:6]

    from tasks.models import Task
    from django.utils import timezone
    my_tasks = Task.objects.filter(assignee=request.user).exclude(status='done')
    overdue_tasks = my_tasks.filter(due_date__lt=timezone.now().date())

    context = {
        'user_orgs': user_orgs,
        'user_projects': user_projects,
        'my_tasks_count': my_tasks.count(),
        'overdue_count': overdue_tasks.count(),
    }
    return render(request, 'organizations/dashboard.html', context)


@login_required
def org_list(request):
    memberships = OrganizationUser.objects.filter(user=request.user).select_related('organization')
    return render(request, 'organizations/org_list.html', {'memberships': memberships})


@login_required
def org_create(request):
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        if not name:
            messages.error(request, 'El nombre es obligatorio.')
        else:
            org = Organization.objects.create(
                name=name,
                logo=request.FILES.get('logo'),
                crm_company_id=request.POST.get('crm_company_id', ''),
            )
            OrganizationUser.objects.create(organization=org, user=request.user, role='owner')
            messages.success(request, f'Organización "{org.name}" creada.')
            return redirect('org_detail', pk=org.pk)
    return render(request, 'organizations/org_form.html', {'title': 'Nueva Organización'})


@login_required
@org_permission(can_manage_members)
def org_edit(request, pk, org=None, org_membership=None):
    now = timezone.now()
    name_locked = False
    name_days_remaining = 0

    if org.name_changed_at:
        delta = now - org.name_changed_at
        if delta < timedelta(days=NAME_CHANGE_DAYS):
            name_locked = True
            name_days_remaining = NAME_CHANGE_DAYS - delta.days

    if request.method == 'POST':
        new_name = request.POST.get('name', '').strip()
        if not new_name:
            messages.error(request, 'El nombre es obligatorio.')
        else:
            name_changed = new_name != org.name
            if name_changed and name_locked:
                messages.error(
                    request,
                    f'El nombre de la organización solo se puede cambiar una vez cada {NAME_CHANGE_DAYS} días. '
                    f'Podrás cambiarlo en {name_days_remaining} día{"s" if name_days_remaining != 1 else ""}.'
                )
            else:
                update_fields = ['crm_company_id', 'updated_at']
                org.crm_company_id = request.POST.get('crm_company_id', '')
                if 'logo' in request.FILES:
                    org.logo = request.FILES['logo']
                    update_fields.append('logo')
                if name_changed:
                    org.name = new_name
                    org.name_changed_at = now
                    update_fields += ['name', 'name_changed_at']
                org.save(update_fields=update_fields)
                messages.success(request, f'Organización "{org.name}" actualizada.')
                return redirect('org_detail', pk=org.pk)

    return render(request, 'organizations/org_form.html', {
        'title': 'Editar Organización',
        'org': org,
        'name_locked': name_locked,
        'name_days_remaining': name_days_remaining,
    })


@login_required
@require_org_member()
def org_detail(request, pk, org=None, org_membership=None):
    members = org.get_active_members()
    projects = Project.objects.filter(organization=org, deleted_at__isnull=True).select_related('owner')
    return render(request, 'organizations/org_detail.html', {
        'org': org,
        'membership': org_membership,
        'members': members,
        'projects': projects,
    })


@login_required
@org_permission(can_manage_members)
def org_invite(request, pk, org=None, org_membership=None):
    if request.method == 'POST':
        email = request.POST.get('email', '').strip()
        role = request.POST.get('role', 'member')
        try:
            user = User.objects.get(email=email)
            _, created = OrganizationUser.objects.get_or_create(
                organization=org, user=user,
                defaults={'role': role}
            )
            if created:
                messages.success(request, f'{user.name} añadido/a a la organización.')
                if user != request.user:
                    from notifications.models import Notification
                    Notification.objects.create(
                        user=user,
                        type='org_added',
                        message=f'{request.user.name} te ha añadido a la organización "{org.name}".',
                    )
            else:
                messages.warning(request, f'{user.name} ya es miembro.')
        except User.DoesNotExist:
            messages.error(request, f'No se encontró usuario con email {email}.')
    return redirect('org_detail', pk=pk)


@login_required
@org_permission(can_manage_members)
def org_member_remove(request, pk, member_pk, org=None, org_membership=None):
    member = get_object_or_404(OrganizationUser, pk=member_pk, organization=org)
    if request.method == 'POST' and member.user != request.user:
        if member.role == 'owner' and is_last_owner(org):
            messages.error(request, 'No puedes eliminar al único propietario de la organización.')
            return redirect('org_detail', pk=pk)
        name = member.user.name
        member.delete()
        messages.success(request, f'{name} eliminado de la organización.')
    return redirect('org_detail', pk=pk)


@login_required
@org_permission(can_manage_members)
def org_member_update(request, pk, member_pk, org=None, org_membership=None):
    member = get_object_or_404(OrganizationUser, pk=member_pk, organization=org)
    if request.method == 'POST':
        role = request.POST.get('role')
        if role in ('member', 'admin', 'owner'):
            if not can_assign_role(org_membership, role):
                messages.error(request, 'Solo el propietario puede asignar el rol de propietario.')
                return redirect('org_detail', pk=pk)
            member.role = role
            member.save()
            messages.success(request, f'Rol de {member.user.name} actualizado.')
    return redirect('org_detail', pk=pk)


@login_required
@org_permission(can_manage_members)
def org_rename_project(request, pk, project_pk, org=None, org_membership=None):
    from django.views.decorators.http import require_POST as _require_post
    project = get_object_or_404(Project, pk=project_pk, organization=org, deleted_at__isnull=True)
    if request.method == 'POST':
        new_name = request.POST.get('name', '').strip()
        if not new_name:
            messages.error(request, 'El nombre del proyecto no puede estar vacío.')
        else:
            project.name = new_name
            project.save(update_fields=['name', 'updated_at'])
            messages.success(request, f'Proyecto renombrado a "{new_name}".')
    return redirect('org_detail', pk=pk)


def _can_view_org_hours(membership):
    return membership is not None and membership.role in ('owner', 'admin')


@login_required
@org_permission(_can_view_org_hours)
def org_hours_overview(request, pk, org=None, org_membership=None):
    from tasks.models import TimeLog

    projects = Project.objects.filter(
        organization=org, deleted_at__isnull=True
    ).prefetch_related('time_logs')

    # Per-project stats
    per_project = []
    total_budget = 0
    total_consumed = 0
    for project in projects:
        consumed = project.time_logs.aggregate(total=Sum('hours'))['total'] or 0
        budget = project.hour_budget
        if budget:
            total_budget += float(budget)
        total_consumed += float(consumed)
        remaining = float(budget) - float(consumed) if budget else None
        pct = min(round(float(consumed) / float(budget) * 100), 100) if budget else 0
        per_project.append({
            'project': project,
            'budget': budget,
            'consumed': consumed,
            'remaining': remaining,
            'pct': pct,
        })
    per_project.sort(key=lambda r: float(r['consumed']), reverse=True)

    # Per-member stats across all org projects
    per_member = (
        TimeLog.objects
        .filter(project__organization=org, project__deleted_at__isnull=True)
        .values('user__id', 'user__name', 'user__avatar')
        .annotate(logged=Sum('hours'))
        .order_by('-logged')
    )

    # Per-member per-project breakdown grouped for template
    member_project_breakdown = (
        TimeLog.objects
        .filter(project__organization=org, project__deleted_at__isnull=True)
        .values('user__id', 'user__name', 'project__id', 'project__name', 'project__key')
        .annotate(logged=Sum('hours'))
        .order_by('user__name', '-logged')
    )
    breakdown_by_user = {}
    for row in member_project_breakdown:
        uid = str(row['user__id'])
        breakdown_by_user.setdefault(uid, []).append(row)

    # Attach breakdown to per_member rows for template iteration
    per_member_with_breakdown = []
    for row in per_member:
        uid = str(row['user__id'])
        per_member_with_breakdown.append({
            'uid': uid,
            'name': row['user__name'],
            'logged': row['logged'],
            'projects': breakdown_by_user.get(uid, []),
        })

    all_projects = list(projects)
    all_members = org.get_active_members()

    return render(request, 'organizations/org_hours_overview.html', {
        'org': org,
        'membership': org_membership,
        'per_project': per_project,
        'per_member': per_member_with_breakdown,
        'total_budget': total_budget or None,
        'total_consumed': total_consumed,
        'total_remaining': (total_budget - total_consumed) if total_budget else None,
        'total_pct': min(round(total_consumed / total_budget * 100), 100) if total_budget else 0,
        'all_projects': all_projects,
        'all_members': all_members,
    })


@login_required
@org_permission(_can_view_org_hours)
def org_hours_export(request, pk, org=None, org_membership=None):
    import csv
    import io
    from datetime import date
    from tasks.models import TimeLog
    from projects.models import Project

    fmt = request.GET.get('format', 'csv')
    date_from = request.GET.get('date_from') or None
    date_to = request.GET.get('date_to') or None
    project_id = request.GET.get('project') or None
    user_id = request.GET.get('user') or None
    task_id = request.GET.get('task') or None

    qs = (
        TimeLog.objects
        .filter(project__organization=org, project__deleted_at__isnull=True)
        .select_related('task', 'project', 'user')
        .order_by('logged_date', 'project__name', 'user__name')
    )
    if date_from:
        qs = qs.filter(logged_date__gte=date_from)
    if date_to:
        qs = qs.filter(logged_date__lte=date_to)
    if project_id:
        qs = qs.filter(project_id=project_id)
    if user_id:
        qs = qs.filter(user_id=user_id)
    if task_id:
        qs = qs.filter(task_id=task_id)

    filename_base = f'horas_{org.name.replace(" ", "_")}_{date.today()}'
    headers = ['Fecha', 'Proyecto', 'Clave', 'Tarea', 'Ref. tarea', 'Usuario', 'Horas', 'Nota']

    def row_data(log):
        return [
            log.logged_date.strftime('%Y-%m-%d'),
            log.project.name,
            log.project.key,
            log.task.title,
            f'{log.project.key}-{log.task.project_sequence}',
            log.user.name,
            float(log.hours),
            log.note or '',
        ]

    if fmt == 'xlsx':
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        from django.http import HttpResponse

        wb = Workbook()
        ws = wb.active
        ws.title = 'Horas'

        # Header row
        ws.append(headers)
        header_fill = PatternFill('solid', fgColor='111111')
        header_font = Font(bold=True, color='FFFFFF')
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')

        for log in qs:
            ws.append(row_data(log))

        # Auto column width
        for col in ws.columns:
            max_len = max(len(str(c.value or '')) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 60)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        response = HttpResponse(
            buf.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename_base}.xlsx"'
        return response

    # Default: CSV
    from django.http import HttpResponse
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename_base}.csv"'
    response.write('﻿')  # BOM for Excel UTF-8 compat
    writer = csv.writer(response)
    writer.writerow(headers)
    for log in qs:
        writer.writerow(row_data(log))
    return response


@login_required
@org_permission(_can_view_org_hours)
def org_hours_monthly(request, pk, org=None, org_membership=None):
    from tasks.models import TimeLog
    from django.db.models.functions import TruncMonth
    from django.db.models import Sum
    from datetime import date

    current_year = date.today().year
    year = int(request.GET.get('year', current_year))

    rows = (
        TimeLog.objects
        .filter(
            project__organization=org,
            project__deleted_at__isnull=True,
            logged_date__year=year,
        )
        .annotate(month=TruncMonth('logged_date'))
        .values('user__id', 'user__name', 'month')
        .annotate(logged=Sum('hours'))
        .order_by('user__name', 'month')
    )

    # Build pivot: {user_id: {month_num: hours}}
    MONTH_NAMES = ['Ene', 'Feb', 'Mar', 'Abr', 'May', 'Jun',
                   'Jul', 'Ago', 'Sep', 'Oct', 'Nov', 'Dic']

    users_map = {}   # uid -> {name, months: {1..12: hours}, total}
    active_months = set()

    for row in rows:
        uid = str(row['user__id'])
        month_num = row['month'].month
        active_months.add(month_num)
        if uid not in users_map:
            users_map[uid] = {'name': row['user__name'], 'months': {}, 'total': 0}
        users_map[uid]['months'][month_num] = float(row['logged'])
        users_map[uid]['total'] += float(row['logged'])

    months = sorted(active_months)
    month_labels = [MONTH_NAMES[m - 1] for m in months]

    # Month totals
    month_totals = {m: 0.0 for m in months}
    for u in users_map.values():
        for m, h in u['months'].items():
            month_totals[m] += h

    # Build flat list for template
    pivot_rows = [
        {
            'name': u['name'],
            'cells': [u['months'].get(m, None) for m in months],
            'total': u['total'],
        }
        for u in sorted(users_map.values(), key=lambda x: x['name'])
    ]

    grand_total = sum(u['total'] for u in users_map.values())

    return render(request, 'organizations/org_hours_monthly.html', {
        'org': org,
        'membership': org_membership,
        'year': year,
        'prev_year': year - 1,
        'next_year': year + 1,
        'months': months,
        'month_labels': month_labels,
        'month_totals': [month_totals[m] for m in months],
        'pivot_rows': pivot_rows,
        'grand_total': grand_total,
        'has_data': bool(pivot_rows),
    })
